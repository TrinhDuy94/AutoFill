from selenium import webdriver
import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import time

global mtime
def read_data():
	wb = load_workbook('/Volumes/DATA/Fan/Datarun.xlsx')
	sheet = wb['Sheet1']

	# List to hold dictionaries
	data = []

	# Iterate through each row in worksheet and fetch values into dict
	for row in islice(sheet.values, 1, sheet.max_row):
	    obj = OrderedDict()
	    obj["ID"] = row[0]
	    obj["Cau1"] = str(row[1])
	    obj["Cau2"] = str(row[2])
	    obj["Cau3"] = str(row[3])
	    obj["Cau4"] = str(row[4])
	    obj["Cau5"] = str(row[5])
	    obj["Cau6"] = str(row[6])
	    obj["Cau7"] = str(row[7])
	    obj["Cau8"] = str(row[8])
	    obj["Cau9"] = str(row[9])
	    obj["Cau10"] = str(row[10])
	    obj["Cau11"] = str(row[11])
	    obj["Cau12"] = str(row[12])
	    obj["Cau13"] = str(row[13])
	    obj["Cau14"] = str(row[14])
	    obj["Cau15"] = str(row[15])
	    obj["Cau16"] = str(row[16])
	    obj["Cau17"] = str(row[17])
	    obj["time"] = str(row[18])
	    data.append(obj)
	return data

def run(x):
	global mtime
	data = read_data()
	driver = webdriver.Firefox(executable_path='/Users/vladimirtrinh/Downloads/geckodriver')
	driver.get('https://docs.google.com/forms/d/e/1FAIpQLSdLxRwv2bramQWG6hQRdeXhXRAppVwG7buHa5uLSLbVIUYpWw/viewform?fbclid=IwAR2KjRsgVqx9bY94HXsmP6KqCrV75HHoLh72aq-PwVcmivDhGwyR99pTdMc')
	for i in data:
		if i["ID"]==x:
			c1 = driver.find_element_by_xpath(i["Cau1"])
			c1.click()
			c2 = driver.find_element_by_xpath(i["Cau2"])
			c2.click()
			c3 = driver.find_element_by_xpath(i["Cau3"])
			c3.click()
			c4 = driver.find_element_by_xpath(i["Cau4"])
			c4.click()
			c5 = driver.find_element_by_xpath(i["Cau5"])
			c5.click()
			c6 = driver.find_element_by_xpath(i["Cau6"])
			c6.click()
			c7 = driver.find_element_by_xpath(i["Cau7"])
			c7.click()
			c8 = driver.find_element_by_xpath(i["Cau8"])
			c8.click()
			c9 = driver.find_element_by_xpath(i["Cau9"])
			c9.click()
			c10 = driver.find_element_by_xpath(i["Cau10"])
			c10.click()
			c11 = driver.find_element_by_xpath(i["Cau11"])
			c11.click()
			c12 = driver.find_element_by_xpath(i["Cau12"])
			c12.click()
			c13 = driver.find_element_by_xpath(i["Cau13"])
			c13.click()
			c14 = driver.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[14]/div/div/div[2]/div/div[1]/div/div[1]/input')
			c14.send_keys(i["Cau14"])
			c15 = driver.find_element_by_xpath(i["Cau15"])
			c15.click()
			c16 = driver.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[16]/div/div/div[2]/div/div[1]/div[2]/textarea')
			c16.send_keys(i["Cau16"])
			c17 = driver.find_element_by_xpath(i["Cau17"])
			c17.click()
			mtime = i["time"]
			submit = driver.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div/div/span/span')
			submit.click()
			time.sleep(2)
	driver.quit()	

def main():
	global mtime
	for i in range(22,30):
		run(i)
		time.sleep(mtime)
		print("finish: "+ i)


if __name__ == '__main__':
	main()
