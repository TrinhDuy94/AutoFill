from selenium import webdriver
import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import time

def read_data():
	wb = load_workbook('/Volumes/DATA/datarun.xlsx')
	sheet = wb['Sheet1']

	# List to hold dictionaries
	data = []

	# Iterate through each row in worksheet and fetch values into dict
	for row in islice(sheet.values, 1, sheet.max_row):
	    obj = OrderedDict()
	    obj["Cau1"] = str(row[0])
	    obj["Cau2"] = str(row[1])
	    obj["Cau3"] = str(row[2])
	    obj["Cau4"] = str(row[3])
	    obj["Cau5"] = str(row[4])
	    obj["Cau6"] = str(row[5])
	    obj["Cau7a"] = str(row[6])
	    obj["Cau7b"] = str(row[7])
	    obj["Cau7c"] = str(row[8])
	    obj["Cau7d"] = str(row[9])
	    obj["Cau8"] = str(row[10])
	    obj["Cau9"] = str(row[11])
	    obj["Cau10"] = str(row[12])
	    obj["Cau11"] = str(row[13])
	    obj["Cau12"] = str(row[14])
	    obj["Cau13"] = str(row[15])
	    data.append(obj)
	return data

def run(x):
	data = read_data()
	driver = webdriver.Chrome()
	driver.get('https://docs.google.com/forms/d/e/1FAIpQLSeEE3Ea11WJvjAjz9_6O5ydNwZcLL9IR45yV7_PNvSn3zFLcg/viewform')
	c1 = driver.find_element_by_xpath(data[x]["Cau1"])
	c1.click()
	c2 = driver.find_element_by_xpath(data[x]["Cau2"])
	c2.click()
	c3 = driver.find_element_by_xpath(data[x]["Cau3"])
	c3.click()		
	c4 = driver.find_element_by_xpath(data[x]["Cau4"])
	c4.click()
	c5 = driver.find_element_by_xpath(data[x]["Cau5"])
	c5.click()
	c6 = driver.find_element_by_xpath(data[x]["Cau6"])
	c6.click()
	c7a = driver.find_element_by_xpath(data[x]["Cau7a"])
	c7a.click()
	c7b = driver.find_element_by_xpath(data[x]["Cau7b"])
	c7b.click()
	c7c = driver.find_element_by_xpath(data[x]["Cau7c"])
	c7c.click()
	c7d = driver.find_element_by_xpath(data[x]["Cau7d"])
	c7d.click()
	c8 = driver.find_element_by_xpath(data[x]["Cau8"])
	c8.click()
	c9 = driver.find_element_by_xpath(data[x]["Cau9"])
	c9.click()
	c10 = driver.find_element_by_xpath(data[x]["Cau10"])
	c10.click()
	c11 = driver.find_element_by_xpath(data[x]["Cau11"])
	c11.click()
	c12 = driver.find_element_by_xpath(data[x]["Cau12"])
	c12.click()
	c13 = driver.find_element_by_xpath(data[x]["Cau13"])
	c13.click()
	submit = driver.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div/div/span/span')
	submit.click()
	time.sleep(1)
	driver.quit()
	
	"""for i in data:
		c1 = driver.find_element_by_xpath(i["Cau1"])
		c1.click()
		c2 = driver.find_element_by_xpath(i["Cau2"])
		c2.click()
		c3 = driver.find_element_by_xpath(i["Cau3"])
		c3.click()
		c4 = driver.find_element_by_xpath(i["Cau4"])
		c4.click()
		c5 = driver.find_element_by_xpath(i["Cau5"])
		c5.click()
		c6 = driver.find_element_by_xpath(i["Cau6"])
		c6.click()
		c7a = driver.find_element_by_xpath(i["Cau7a"])
		c7a.click()
		c7b = driver.find_element_by_xpath(i["Cau7b"])
		c7b.click()
		c7c = driver.find_element_by_xpath(i["Cau7c"])
		c7c.click()
		c7d = driver.find_element_by_xpath(i["Cau7d"])
		c7d.click()
		c8 = driver.find_element_by_xpath(i["Cau8"])
		c8.click()
		c9 = driver.find_element_by_xpath(i["Cau9"])
		c9.click()
		c10 = driver.find_element_by_xpath(i["Cau10"])
		c10.click()
		c11 = driver.find_element_by_xpath(i["Cau11"])
		c11.click()
		c12 = driver.find_element_by_xpath(i["Cau12"])
		c12.click()
		c13 = driver.find_element_by_xpath(i["Cau13"])
		c13.click()
		submit = driver.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div/div/span/span')
		submit.click()
		time.sleep(1)
	driver.quit()"""

def main():
	for i in range (1,84):
		run(i)

if __name__ == '__main__':
	main()
